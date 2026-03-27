#!/usr/bin/env python3
"""
Export DTRules decision tables from XML to Excel spreadsheet format.

This script parses the TaxReturn_dt.xml and creates an Excel workbook with
one sheet per decision table, formatted in the classic decision table grid layout.
"""

import xml.etree.ElementTree as ET
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter


def parse_decision_tables(xml_path):
    """Parse decision tables from XML file."""
    tree = ET.parse(xml_path)
    root = tree.getroot()

    tables = []
    for dt in root.findall('decision_table'):
        table = {
            'name': dt.findtext('table_name', ''),
            'type': '',
            'comments': '',
            'table_number': '',
            'contexts': [],
            'initial_actions': [],
            'conditions': [],
            'actions': [],
            'policy_statements': [],
            'max_column': 0
        }

        # Parse attribute fields
        attr_fields = dt.find('attribute_fields')
        if attr_fields is not None:
            table['type'] = attr_fields.findtext('Type', '')
            table['comments'] = attr_fields.findtext('COMMENTS', '')
            table['table_number'] = attr_fields.findtext('TABLE_NUMBER', '')

        # Parse contexts
        contexts = dt.find('contexts')
        if contexts is not None:
            for ctx in contexts.findall('context'):
                table['contexts'].append({
                    'description': ctx.findtext('context_description', ''),
                    'postfix': ctx.findtext('context_postfix', '')
                })

        # Parse initial actions
        initial_actions = dt.find('initial_actions')
        if initial_actions is not None:
            for ia in initial_actions.findall('initial_action'):
                table['initial_actions'].append({
                    'description': ia.findtext('action_description', ''),
                    'postfix': ia.findtext('action_postfix', '')
                })

        # Parse conditions
        conditions = dt.find('conditions')
        if conditions is not None:
            for cond in conditions.findall('condition_details'):
                condition = {
                    'number': cond.findtext('condition_number', ''),
                    'comment': cond.findtext('condition_comment', ''),
                    'description': cond.findtext('condition_description', ''),
                    'postfix': cond.findtext('condition_postfix', ''),
                    'columns': {}
                }
                for col in cond.findall('condition_column'):
                    col_num = int(col.get('column_number', 0))
                    col_val = col.get('column_value', '')
                    condition['columns'][col_num] = col_val
                    table['max_column'] = max(table['max_column'], col_num)
                table['conditions'].append(condition)

        # Parse actions
        actions = dt.find('actions')
        if actions is not None:
            for act in actions.findall('action_details'):
                action = {
                    'number': act.findtext('action_number', ''),
                    'comment': act.findtext('action_comment', ''),
                    'description': act.findtext('action_description', ''),
                    'postfix': act.findtext('action_postfix', ''),
                    'columns': {}
                }
                for col in act.findall('action_column'):
                    col_num = int(col.get('column_number', 0))
                    col_val = col.get('column_value', '')
                    action['columns'][col_num] = col_val
                    table['max_column'] = max(table['max_column'], col_num)
                table['actions'].append(action)

        # Parse policy statements
        policy = dt.find('policy_statements')
        if policy is not None:
            for ps in policy.findall('policy_statement'):
                table['policy_statements'].append({
                    'column': ps.get('column', ''),
                    'description': ps.findtext('policy_description', ''),
                    'postfix': ps.findtext('policy_statement_postfix', '')
                })

        tables.append(table)

    return tables


def create_excel(tables, output_path):
    """Create Excel workbook from parsed decision tables."""
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")

    condition_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    action_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Create index sheet
    idx_ws = wb.create_sheet("Index")
    idx_ws['A1'] = "Decision Table Index"
    idx_ws['A1'].font = Font(bold=True, size=14)
    idx_ws['A3'] = "Table #"
    idx_ws['B3'] = "Table Name"
    idx_ws['C3'] = "Type"
    idx_ws['D3'] = "Description"
    for col in ['A', 'B', 'C', 'D']:
        idx_ws[f'{col}3'].font = header_font
        idx_ws[f'{col}3'].fill = header_fill
        idx_ws[f'{col}3'].font = header_font_white

    idx_row = 4
    for table in tables:
        idx_ws[f'A{idx_row}'] = table['table_number']
        idx_ws[f'B{idx_row}'] = table['name']
        idx_ws[f'C{idx_row}'] = table['type']
        idx_ws[f'D{idx_row}'] = table['comments'][:100] + "..." if len(table['comments']) > 100 else table['comments']
        idx_row += 1

    idx_ws.column_dimensions['A'].width = 10
    idx_ws.column_dimensions['B'].width = 35
    idx_ws.column_dimensions['C'].width = 10
    idx_ws.column_dimensions['D'].width = 60

    # Create sheet for each table
    for table in tables:
        # Truncate sheet name to 31 chars (Excel limit)
        sheet_name = table['name'][:31]
        ws = wb.create_sheet(sheet_name)

        row = 1

        # Table header
        ws[f'A{row}'] = f"Decision Table: {table['name']}"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1

        # Table metadata
        ws[f'A{row}'] = f"Table Number: {table['table_number']}"
        ws[f'C{row}'] = f"Type: {table['type']}"
        row += 1

        if table['comments']:
            ws[f'A{row}'] = f"Description: {table['comments']}"
            ws[f'A{row}'].alignment = left_align
            ws.merge_cells(f'A{row}:H{row}')
            row += 1

        row += 1  # Blank row

        # Initial actions (if any)
        if table['initial_actions']:
            ws[f'A{row}'] = "Initial Actions"
            ws[f'A{row}'].font = header_font
            row += 1
            for ia in table['initial_actions']:
                ws[f'A{row}'] = ia['description']
                ws[f'B{row}'] = ia['postfix'].strip()
                row += 1
            row += 1

        # Contexts (if any)
        if table['contexts']:
            ws[f'A{row}'] = "Contexts"
            ws[f'A{row}'].font = header_font
            row += 1
            for ctx in table['contexts']:
                ws[f'A{row}'] = ctx['description']
                ws[f'B{row}'] = ctx['postfix'].strip()
                row += 1
            row += 1

        # Decision table grid
        grid_start_row = row
        num_cols = table['max_column']

        # Column headers
        ws[f'A{row}'] = "#"
        ws[f'B{row}'] = "Description"
        ws[f'C{row}'] = "Comment"

        col_offset = 4  # D column starts rule columns
        for col_num in range(1, num_cols + 1):
            col_letter = get_column_letter(col_offset + col_num - 1)
            ws[f'{col_letter}{row}'] = f"Col {col_num}"
            ws[f'{col_letter}{row}'].font = header_font_white
            ws[f'{col_letter}{row}'].fill = header_fill
            ws[f'{col_letter}{row}'].alignment = center_align
            ws[f'{col_letter}{row}'].border = thin_border

        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = header_font_white
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].border = thin_border

        row += 1

        # Conditions section header
        ws[f'A{row}'] = "CONDITIONS"
        ws[f'A{row}'].font = Font(bold=True, italic=True)
        ws.merge_cells(f'A{row}:C{row}')
        row += 1

        # Conditions
        for cond in table['conditions']:
            ws[f'A{row}'] = cond['number']
            ws[f'B{row}'] = cond['description']
            ws[f'C{row}'] = cond['comment']

            ws[f'A{row}'].fill = condition_fill
            ws[f'B{row}'].fill = condition_fill
            ws[f'C{row}'].fill = condition_fill
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].border = thin_border
            ws[f'C{row}'].border = thin_border
            ws[f'B{row}'].alignment = left_align

            for col_num in range(1, num_cols + 1):
                col_letter = get_column_letter(col_offset + col_num - 1)
                val = cond['columns'].get(col_num, '')
                ws[f'{col_letter}{row}'] = val
                ws[f'{col_letter}{row}'].alignment = center_align
                ws[f'{col_letter}{row}'].border = thin_border
                ws[f'{col_letter}{row}'].fill = condition_fill

            row += 1

        # Actions section header
        ws[f'A{row}'] = "ACTIONS"
        ws[f'A{row}'].font = Font(bold=True, italic=True)
        ws.merge_cells(f'A{row}:C{row}')
        row += 1

        # Actions
        for act in table['actions']:
            ws[f'A{row}'] = act['number']
            ws[f'B{row}'] = act['description']
            ws[f'C{row}'] = act['comment']

            ws[f'A{row}'].fill = action_fill
            ws[f'B{row}'].fill = action_fill
            ws[f'C{row}'].fill = action_fill
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].border = thin_border
            ws[f'C{row}'].border = thin_border
            ws[f'B{row}'].alignment = left_align

            for col_num in range(1, num_cols + 1):
                col_letter = get_column_letter(col_offset + col_num - 1)
                val = act['columns'].get(col_num, '')
                ws[f'{col_letter}{row}'] = val
                ws[f'{col_letter}{row}'].alignment = center_align
                ws[f'{col_letter}{row}'].border = thin_border
                ws[f'{col_letter}{row}'].fill = action_fill

            row += 1

        # Policy statements (if any)
        if table['policy_statements']:
            row += 1
            ws[f'A{row}'] = "Policy Statements"
            ws[f'A{row}'].font = header_font
            row += 1
            for ps in table['policy_statements']:
                ws[f'A{row}'] = f"Column {ps['column']}"
                ws[f'B{row}'] = ps['description']
                row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 35
        for col_num in range(1, num_cols + 1):
            col_letter = get_column_letter(col_offset + col_num - 1)
            ws.column_dimensions[col_letter].width = 8

    # Save workbook
    wb.save(output_path)
    print(f"Exported {len(tables)} decision tables to: {output_path}")


def main():
    # Paths
    script_dir = Path(__file__).parent
    xml_path = script_dir / "xml" / "TaxReturn_dt.xml"
    output_path = script_dir / "TaxReturn_DecisionTables.xlsx"

    if not xml_path.exists():
        print(f"Error: XML file not found: {xml_path}")
        sys.exit(1)

    print(f"Parsing decision tables from: {xml_path}")
    tables = parse_decision_tables(xml_path)
    print(f"Found {len(tables)} decision tables")

    create_excel(tables, output_path)


if __name__ == "__main__":
    main()
